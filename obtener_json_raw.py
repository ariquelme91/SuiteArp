"""Script para obtener JSON RAW de Buk y guardarlo en Excel"""
import os
import json
import requests
from dotenv import load_dotenv
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

load_dotenv()

token = os.getenv('BUK_API_TOKEN')
subdomain = os.getenv('BUK_SUBDOMAIN')

base_url = f"https://{subdomain}.buk.cl/api/v1/chile"
headers = {
    "auth_token": token,
    "Content-Type": "application/json"
}

ruts = ["17771319-8"]

def format_rut(rut):
    return re.sub(r'[.-]', '', rut).upper()

def flatten_dict(d, parent_key='', sep='_'):
    """Flattena diccionarios anidados"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            items.append((new_key, json.dumps(v, indent=2, default=str, ensure_ascii=False)))
        else:
            items.append((new_key, v))
    return dict(items)

print("Obteniendo JSON de empleados y generando Excel...")

wb = Workbook()
wb.remove(wb.active)

for rut in ruts:
    print(f"\nProcesando: {rut}")

    try:
        rut_formatted = format_rut(rut)
        response = requests.get(
            f"{base_url}/employees/{rut_formatted}",
            headers=headers,
            timeout=10
        )
        response.raise_for_status()

        data = response.json()

        # Extraer datos del empleado
        if "data" in data:
            employee_data = data["data"]
        else:
            employee_data = data

        nombre = employee_data.get('full_name', rut)
        print(f"  ✓ {nombre}")

        # ===== HOJA 1: JSON FLATTENED (TODOS LOS CAMPOS) =====
        ws_flat = wb.create_sheet("JSON Completo")

        flat_data = flatten_dict(employee_data)

        # Headers
        ws_flat['A1'] = "Campo"
        ws_flat['B1'] = "Valor"
        ws_flat['A1'].font = Font(bold=True, color="FFFFFF")
        ws_flat['B1'].font = Font(bold=True, color="FFFFFF")
        ws_flat['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws_flat['B1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        # Datos
        row = 2
        for key, value in flat_data.items():
            ws_flat[f'A{row}'] = key
            ws_flat[f'B{row}'] = str(value) if value is not None else ""
            row += 1

        ws_flat.column_dimensions['A'].width = 50
        ws_flat.column_dimensions['B'].width = 80
        ws_flat.column_dimensions['B'].alignment = Alignment(wrap_text=True)

        # ===== HOJA 2: JSON RAW (formateado) =====
        ws_raw = wb.create_sheet("JSON Raw")
        json_str = json.dumps(employee_data, indent=2, default=str, ensure_ascii=False)
        ws_raw['A1'] = json_str
        ws_raw.column_dimensions['A'].width = 150
        ws_raw.column_dimensions['A'].alignment = Alignment(wrap_text=True)

        # ===== HOJA 3: CURRENT JOB DETALLADO =====
        if "current_job" in employee_data:
            ws_job = wb.create_sheet("Current Job")
            current_job = employee_data["current_job"]

            flat_job = flatten_dict(current_job)

            ws_job['A1'] = "Campo"
            ws_job['B1'] = "Valor"
            ws_job['A1'].font = Font(bold=True, color="FFFFFF")
            ws_job['B1'].font = Font(bold=True, color="FFFFFF")
            ws_job['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            ws_job['B1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

            row = 2
            for key, value in flat_job.items():
                ws_job[f'A{row}'] = key
                ws_job[f'B{row}'] = str(value) if value is not None else ""
                row += 1

            ws_job.column_dimensions['A'].width = 50
            ws_job.column_dimensions['B'].width = 80

        # ===== HOJA 4: BÚSQUEDA DE HABERES Y DESCUENTOS =====
        ws_search = wb.create_sheet("Busqueda Items")
        ws_search['A1'] = "Tipo"
        ws_search['B1'] = "Clave"
        ws_search['C1'] = "Contenido"

        for col in ['A', 'B', 'C']:
            ws_search[f'{col}1'].font = Font(bold=True, color="FFFFFF")
            ws_search[f'{col}1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        row = 2
        # Buscar claves que contengan "item", "bono", "haberes", "descuento"
        keywords = ["fixed_item", "bono", "haber", "descuento", "allowance", "deduction", "benefit"]

        for key, value in flat_data.items():
            key_lower = key.lower()
            if any(kw in key_lower for kw in keywords):
                ws_search[f'A{row}'] = key
                ws_search[f'B{row}'] = key
                ws_search[f'C{row}'] = str(value)[:500] if value else ""
                row += 1

        ws_search.column_dimensions['A'].width = 40
        ws_search.column_dimensions['B'].width = 40
        ws_search.column_dimensions['C'].width = 100

    except Exception as e:
        print(f"  ERROR: {str(e)}")
        import traceback
        traceback.print_exc()

# Guardar archivo Excel
filename = f"JSON_Raw_Empleados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(filename)

print(f"\n✅ Archivo Excel creado: {filename}")
print(f"   📁 Ubicación: C:\\Users\\ariquelme\\propuestas de renta\\{filename}")
print(f"\n   Hojas creadas:")
print(f"   - JSON Completo: Todos los campos del empleado")
print(f"   - JSON Raw: JSON formateado para revisión completa")
print(f"   - Current Job: Detalles de la ficha vigente")
print(f"   - Busqueda Items: Campos relacionados a haberes/descuentos")

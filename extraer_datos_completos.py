"""Script para extraer TODOS los datos de Buk incluyendo Haberes y Descuentos"""
import os
import json
import requests
import re
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

load_dotenv()

token = os.getenv('BUK_API_TOKEN')
subdomain = os.getenv('BUK_SUBDOMAIN')

base_url = f"https://{subdomain}.buk.cl/api/v1/chile"
headers = {
    "auth_token": token,
    "Content-Type": "application/json"
}

ruts = ["17771319-8", "13531650-4", "18.732.184-0"]

def format_rut(rut):
    return re.sub(r'[.-]', '', rut).upper()

def add_header_style(ws, row):
    """Agrega estilo a los headers"""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

print("Extrayendo datos completos de Buk...")

wb = Workbook()
wb.remove(wb.active)

for rut in ruts:
    print(f"\nProcesando: {rut}")

    try:
        rut_formatted = format_rut(rut)
        response = requests.get(
            f"{base_url}/employees/{rut_formatted}",
            headers=headers,
            timeout=10
        )
        response.raise_for_status()

        raw_data = response.json()
        employee_data = raw_data.get("data", raw_data)

        # Verificar que sea ficha vigente/activa
        status = employee_data.get("status", "")
        current_job = employee_data.get("current_job", {})
        job_status = current_job.get("status", "")

        if status != "activo" and job_status != "activo":
            print(f"  SALTADO - No es ficha activa (status: {status})")
            continue

        nombre = employee_data.get("full_name", rut)
        print(f"  OK - {nombre}")

        # ===== HOJA 1: DATOS BÁSICOS =====
        ws_basicos = wb.create_sheet(f"{rut} - Basicos")

        row = 1
        datos_basicos = {
            "RUT": employee_data.get("rut"),
            "Nombre": employee_data.get("full_name"),
            "Email": employee_data.get("email"),
            "Cargo": current_job.get("name"),
            "Empresa": current_job.get("company_name"),
            "Estado": current_job.get("status"),
            "Sueldo Base": current_job.get("base_wage"),
            "Fecha Ingreso": current_job.get("start_date"),
            "Contrato": current_job.get("contract_type"),
            "Fondo Pensión": employee_data.get("pension_fund"),
        }

        for key, value in datos_basicos.items():
            ws_basicos[f"A{row}"] = key
            ws_basicos[f"B{row}"] = value
            ws_basicos[f"A{row}"].font = Font(bold=True)
            row += 1

        ws_basicos.column_dimensions["A"].width = 25
        ws_basicos.column_dimensions["B"].width = 40

        # ===== HOJA 2: HABERES =====
        haberes = current_job.get("fixed_items", [])
        if haberes:
            ws_haberes = wb.create_sheet(f"{rut} - Haberes")

            headers = ["Nombre", "Valor", "Forma de Cálculo", "Incluido en", "Estado", "Asignado en"]
            for col, header in enumerate(headers, 1):
                ws_haberes.cell(row=1, column=col).value = header

            add_header_style(ws_haberes, 1)

            for row, haber in enumerate(haberes, 2):
                ws_haberes.cell(row=row, column=1).value = haber.get("name")
                ws_haberes.cell(row=row, column=2).value = haber.get("value")
                ws_haberes.cell(row=row, column=3).value = haber.get("calculation_formula")
                ws_haberes.cell(row=row, column=4).value = haber.get("included_in")
                ws_haberes.cell(row=row, column=5).value = haber.get("status")
                ws_haberes.cell(row=row, column=6).value = haber.get("assigned_at")

            for col in range(1, len(headers) + 1):
                ws_haberes.column_dimensions[chr(64 + col)].width = 20

        # ===== HOJA 3: DESCUENTOS =====
        discounts = current_job.get("discounts", [])
        if discounts:
            ws_descuentos = wb.create_sheet(f"{rut} - Descuentos")

            headers = ["Nombre", "Valor", "Forma de Cálculo", "Estado", "Asignado en"]
            for col, header in enumerate(headers, 1):
                ws_descuentos.cell(row=1, column=col).value = header

            add_header_style(ws_descuentos, 1)

            for row, descuento in enumerate(discounts, 2):
                ws_descuentos.cell(row=row, column=1).value = descuento.get("name")
                ws_descuentos.cell(row=row, column=2).value = descuento.get("value")
                ws_descuentos.cell(row=row, column=3).value = descuento.get("calculation_formula")
                ws_descuentos.cell(row=row, column=4).value = descuento.get("status")
                ws_descuentos.cell(row=row, column=5).value = descuento.get("assigned_at")

            for col in range(1, len(headers) + 1):
                ws_descuentos.column_dimensions[chr(64 + col)].width = 20

        # ===== HOJA 4: TODOS LOS ATRIBUTOS =====
        ws_todo = wb.create_sheet(f"{rut} - Todos")

        def flatten_dict(d, parent_key='', sep='_'):
            items = []
            for k, v in d.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(flatten_dict(v, new_key, sep=sep).items())
                elif isinstance(v, list):
                    if len(str(v)) > 100:
                        items.append((new_key, f"[{len(v)} items]"))
                    else:
                        items.append((new_key, json.dumps(v, default=str)))
                else:
                    items.append((new_key, v))
            return dict(items)

        flat_data = flatten_dict(employee_data)

        row = 1
        for key, value in flat_data.items():
            ws_todo[f"A{row}"] = key
            ws_todo[f"B{row}"] = str(value) if value is not None else ""
            ws_todo[f"A{row}"].font = Font(bold=True)
            row += 1

        ws_todo.column_dimensions["A"].width = 40
        ws_todo.column_dimensions["B"].width = 60

    except Exception as e:
        print(f"  ERROR: {str(e)}")

# Guardar archivo Excel
filename = f"Data_Buk_Completa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(filename)

print(f"\nArchivo Excel creado: {filename}")
print(f"Ubicacion: C:\\Users\\ariquelme\\propuestas de renta\\{filename}")

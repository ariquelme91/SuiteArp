"""
Exportador de análisis a Excel.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Any, Optional
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exporta análisis a Excel."""

    # Colores
    COLOR_HEADER = "1F4E78"
    COLOR_SECTION = "D9E1F2"
    COLOR_TOTAL = "E2EFDA"

    def export_analysis(
        self,
        analyses: List[Dict[str, Any]],
        metrics: Dict[str, Any],
        output_filename: str,
        empresa: str = "Todas",
        area: str = "Todas",
    ) -> bool:
        """
        Exporta análisis a Excel con múltiples hojas.

        Args:
            analyses: Lista de análisis de empleados
            metrics: Métricas resumidas
            output_filename: Ruta del archivo
            empresa: Empresa filtrada
            area: Área filtrada

        Returns:
            True si fue exitoso
        """
        try:
            from src.analysis.metrics_calculator import AdvancedMetricsCalculator

            wb = Workbook()
            wb.remove(wb.active)

            # Calcular métricas avanzadas
            advanced_metrics = AdvancedMetricsCalculator.calculate_metrics(analyses)

            # Hoja 1: Resumen
            self._add_summary_sheet(wb, metrics, empresa, area)

            # Hoja 2: Métricas Avanzadas
            self._add_advanced_metrics_sheet(wb, advanced_metrics)

            # Hoja 3: Detalle
            self._add_detail_sheet(wb, analyses)

            # Hoja 4: Top Aumentos
            self._add_top_increases_sheet(wb, analyses)

            wb.save(output_filename)
            logger.info(f"Excel generado: {output_filename}")
            return True

        except Exception as e:
            logger.error(f"Error exportando Excel: {e}")
            return False

    def _add_summary_sheet(self, wb, metrics: Dict, empresa: str, area: str):
        """Agrega hoja de resumen."""
        ws = wb.create_sheet("Resumen")

        # Encabezado
        ws.merge_cells("A1:D1")
        title = ws["A1"]
        title.value = "ANÁLISIS DE AUMENTOS DE RENTA"
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
        title.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 25

        # Información del reporte
        row = 3
        ws[f"A{row}"].value = "Fecha Reporte:"
        ws[f"B{row}"].value = datetime.now().strftime("%d/%m/%Y %H:%M")

        row += 1
        ws[f"A{row}"].value = "Empresa:"
        ws[f"B{row}"].value = empresa

        row += 1
        ws[f"A{row}"].value = "Área:"
        ws[f"B{row}"].value = area

        # KPIs
        row = 8
        self._add_section_header(ws, row, "MÉTRICAS CLAVE")
        row += 1

        kpis = [
            ("Total de Empleados", metrics.get("total_empleados", 0)),
            ("Aumento Total Invertido", f"${metrics.get('aumento_total_invertido', 0):,.0f}"),
            ("Aumento Promedio %", f"{metrics.get('aumento_promedio_pct', 0):.1f}%"),
            ("Aumento Promedio $", f"${metrics.get('aumento_promedio_monto', 0):,.0f}"),
        ]

        for label, value in kpis:
            ws[f"A{row}"].value = label
            ws[f"B{row}"].value = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Empleado con mayor aumento
        if metrics.get("empleado_mayor_aumento"):
            row += 1
            self._add_section_header(ws, row, "TOP EMPLEADOS")
            row += 1

            emp = metrics["empleado_mayor_aumento"]
            ws[f"A{row}"].value = "Mayor Aumento:"
            ws[f"B{row}"].value = f"{emp['nombre']} - ${emp['aumento_total']:,.0f}"
            row += 1

            emp = metrics.get("empleado_menor_aumento")
            if emp:
                ws[f"A{row}"].value = "Menor Aumento:"
                ws[f"B{row}"].value = f"{emp['nombre']} - ${emp['aumento_total']:,.0f}"

        # Ancho de columnas
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25

    def _add_detail_sheet(self, wb, analyses: List[Dict]):
        """Agrega hoja con detalle de empleados."""
        ws = wb.create_sheet("Detalle Empleados")

        # Encabezados
        headers = [
            "RUT", "Nombre", "Edad", "Empresa", "Área", "Cargo Actual",
            "Fecha Ingreso", "Meses Empresa", "Meses Puesto", "Sueldo Inicial", "Sueldo Actual",
            "Aumento $", "Aumento %", "Prom. Anual"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Datos
        for row, analysis in enumerate(analyses, 2):
            ws.cell(row=row, column=1).value = analysis.get("rut")
            ws.cell(row=row, column=2).value = analysis.get("nombre")
            ws.cell(row=row, column=3).value = analysis.get("edad")
            ws.cell(row=row, column=4).value = analysis.get("empresa")
            ws.cell(row=row, column=5).value = analysis.get("area", "")
            ws.cell(row=row, column=6).value = analysis.get("cargo_actual")
            ws.cell(row=row, column=7).value = analysis.get("fecha_ingreso")
            ws.cell(row=row, column=8).value = analysis.get("meses_en_empresa")
            ws.cell(row=row, column=9).value = analysis.get("meses_en_puesto")
            ws.cell(row=row, column=10).value = analysis.get("sueldo_inicial")
            ws.cell(row=row, column=11).value = analysis.get("sueldo_actual")
            ws.cell(row=row, column=12).value = analysis.get("aumento_total")
            ws.cell(row=row, column=13).value = analysis.get("aumento_total_pct")
            ws.cell(row=row, column=14).value = analysis.get("promedio_aumento_anual")

            # Formatear números
            for col in [8, 9, 10, 11, 12, 14]:
                ws.cell(row=row, column=col).number_format = "#,##0"
            ws.cell(row=row, column=13).number_format = "0.00%"

        # Ancho de columnas
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15

    def _add_top_increases_sheet(self, wb, analyses: List[Dict]):
        """Agrega hoja con top 20 aumentos."""
        ws = wb.create_sheet("Top 20 Aumentos")

        # Ordenar por aumento total
        sorted_analyses = sorted(
            analyses, key=lambda x: x.get("aumento_total", 0), reverse=True
        )[:20]

        # Encabezados
        headers = ["Ranking", "Nombre", "Cargo", "Sueldo Inicial", "Sueldo Actual", "Aumento $", "Aumento %"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Datos
        for rank, analysis in enumerate(sorted_analyses, 1):
            row = rank + 1
            ws.cell(row=row, column=1).value = rank
            ws.cell(row=row, column=2).value = analysis.get("nombre")
            ws.cell(row=row, column=3).value = analysis.get("cargo_actual")
            ws.cell(row=row, column=4).value = analysis.get("sueldo_inicial")
            ws.cell(row=row, column=5).value = analysis.get("sueldo_actual")
            ws.cell(row=row, column=6).value = analysis.get("aumento_total")
            ws.cell(row=row, column=7).value = analysis.get("aumento_total_pct")

            # Formatear números
            for col in [4, 5, 6]:
                ws.cell(row=row, column=col).number_format = "#,##0"
            ws.cell(row=row, column=7).number_format = "0.00%"

            # Color alternado
            if rank % 2 == 0:
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = PatternFill(
                        start_color="F5F5F5", fill_type="solid"
                    )

        # Ancho de columnas
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 25

    def _add_advanced_metrics_sheet(self, wb, advanced_metrics: Dict):
        """Agrega hoja con métricas avanzadas."""
        ws = wb.create_sheet("Métricas Avanzadas")

        # Encabezado
        ws.merge_cells("A1:D1")
        title = ws["A1"]
        title.value = "MÉTRICAS AVANZADAS DE COMPENSACIONES"
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")
        title.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 25

        # Métricas principales
        row = 3
        metrics_data = [
            ("Impacto en Masa Salarial", f"{advanced_metrics.get('impacto_masa_salarial_pct', 0):.2f}%"),
            ("Costo Anualizado", f"${advanced_metrics.get('costo_anualizado', 0):,.0f}"),
            ("Concentración Top 3", f"{advanced_metrics.get('concentracion_top_3_pct', 0):.1f}%"),
            ("Tasa de Cobertura", f"{advanced_metrics.get('tasa_cobertura_pct', 0):.1f}%"),
        ]

        for label, value in metrics_data:
            ws[f"A{row}"].value = label
            ws[f"B{row}"].value = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Sección de estadísticos
        row += 1
        self._add_section_header(ws, row, "ESTADÍSTICOS - AUMENTO ($)")
        row += 1

        stat_headers = ["Métrica", "Valor"]
        for col, header in enumerate(stat_headers, 1):
            ws.cell(row=row, column=col).value = header
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")

        row += 1
        stat_data_monto = [
            ("Mediana", advanced_metrics.get("mediana_aumento_monto", 0)),
            ("P25", advanced_metrics.get("p25_aumento_monto", 0)),
            ("P75", advanced_metrics.get("p75_aumento_monto", 0)),
            ("Desv. Estándar", advanced_metrics.get("desv_est_aumento_monto", 0)),
        ]

        for label, value in stat_data_monto:
            ws[f"A{row}"].value = label
            ws[f"B{row}"].value = value
            ws[f"B{row}"].number_format = "#,##0"
            row += 1

        # Sección de estadísticos %
        row += 1
        self._add_section_header(ws, row, "ESTADÍSTICOS - AUMENTO (%)")
        row += 1

        for col, header in enumerate(stat_headers, 1):
            ws.cell(row=row, column=col).value = header
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=row, column=col).fill = PatternFill(start_color=self.COLOR_HEADER, fill_type="solid")

        row += 1
        stat_data_pct = [
            ("Mediana", advanced_metrics.get("mediana_aumento_pct", 0)),
            ("P25", advanced_metrics.get("p25_aumento_pct", 0)),
            ("P75", advanced_metrics.get("p75_aumento_pct", 0)),
            ("Desv. Estándar", advanced_metrics.get("desv_est_aumento_pct", 0)),
        ]

        for label, value in stat_data_pct:
            ws[f"A{row}"].value = label
            ws[f"B{row}"].value = value
            ws[f"B{row}"].number_format = "0.00%"
            row += 1

        # Antigüedad del último aumento
        row += 1
        ws[f"A{row}"].value = "Antigüedad Promedio (Meses)"
        ws[f"B{row}"].value = advanced_metrics.get("antigüedad_promedio_meses_ultimo_aumento", 0)
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"].number_format = "0.0"

        # Ancho de columnas
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20

    def _add_section_header(self, ws, row: int, title: str):
        """Agrega encabezado de sección."""
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=self.COLOR_SECTION, fill_type="solid")
        cell.alignment = Alignment(horizontal="left")
        ws.row_dimensions[row].height = 18

"""Generador de reportes Excel con formato ejecutivo."""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Tuple
from .simulator import ComparisonResult
import logging

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exporta comparativas de Propuestas de Renta a Excel (.xlsx)."""

    # Colores de tema
    COLOR_HEADER = "1F4E78"  # Azul oscuro
    COLOR_SECTION = "D9E1F2"  # Azul claro
    COLOR_TOTAL = "E2EFDA"   # Verde claro
    COLOR_POSITIVE = "C6EFCE"  # Verde positivo
    COLOR_NEGATIVE = "FFC7CE"  # Rojo negativo

    def __init__(self):
        """Inicializa exportador."""
        self.workbook = None

    def export_comparison(
        self,
        comparison: ComparisonResult,
        output_filename: str,
        company_name: str = "Empresa",
        prepared_by: str = "RRHH",
        standard_proposals: Dict[int, ComparisonResult] = None,
        current_company: str = "",
        current_position: str = "",
        current_supervisor: str = "",
        proposal_company: str = "",
        proposal_position: str = "",
        proposal_supervisor: str = "",
        salary_history: list = None,
        proposal_reasons: list = None,
    ) -> bool:
        """
        Exporta comparativa a archivo Excel.

        Args:
            comparison: ComparisonResult con datos a exportar
            output_filename: Ruta del archivo Excel
            company_name: Nombre de la empresa
            prepared_by: Persona que prepara el documento
            standard_proposals: Dict {porcentaje: ComparisonResult} para propuestas estándar
            current_company: Empresa actual (Buk)
            current_position: Cargo actual (Buk)
            current_supervisor: Jefe actual (Buk)
            proposal_company: Empresa propuesta
            proposal_position: Cargo propuesto
            proposal_supervisor: Jefe propuesto

        Returns:
            True si es exitoso, False en caso contrario
        """
        try:
            self.workbook = Workbook()
            ws = self.workbook.active
            ws.title = "Propuesta de Renta"

            self._setup_sheet(ws)
            self._add_header(ws, company_name, prepared_by)
            self._add_employee_info(ws, comparison, current_company, current_position, current_supervisor,
                                   proposal_company or current_company, proposal_position or current_position,
                                   proposal_supervisor or current_supervisor)
            self._add_comparison_table(ws, comparison)
            self._add_impact_summary(ws, comparison)
            self._add_observations(ws, proposal_reasons)

            # Agregar propuestas estándar si se proporcionan
            if standard_proposals:
                for percentage in sorted(standard_proposals.keys()):
                    proposal = standard_proposals[percentage]
                    ws_std = self.workbook.create_sheet(f"Propuesta {percentage}%")
                    self._setup_sheet(ws_std)
                    self._add_header(ws_std, company_name, prepared_by)
                    self._add_employee_info(ws_std, proposal, current_company, current_position, current_supervisor,
                                          proposal_company or current_company, proposal_position or current_position,
                                          proposal_supervisor or current_supervisor)
                    self._add_comparison_table(ws_std, proposal)
                    self._add_impact_summary(ws_std, proposal)

            # Agregar historial de sueldos si se proporciona
            if salary_history:
                ws_history = self.workbook.create_sheet("Historial de Sueldos")
                self._add_salary_history(ws_history, salary_history)

            self.workbook.save(output_filename)
            logger.info(f"Archivo Excel generado: {output_filename}")
            return True

        except Exception as e:
            logger.error(f"Error generando Excel: {e}")
            return False

    def _setup_sheet(self, ws):
        """Configura ancho de columnas y configuraciones generales."""
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 22
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18

    def _add_header(self, ws, company_name: str, prepared_by: str):
        """Agrega encabezado con información de empresa."""
        ws.merge_cells("A1:E1")
        title = ws["A1"]
        title.value = "PROPUESTA DE RENTA"
        title.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        ws.merge_cells("A2:E2")
        subtitle = ws["A2"]
        subtitle.value = company_name
        subtitle.font = Font(name="Calibri", size=11, italic=True)
        subtitle.alignment = Alignment(horizontal="center")

        ws["A4"].value = "Fecha de Generación:"
        ws["B4"].value = datetime.now().strftime("%d/%m/%Y %H:%M")
        ws["A5"].value = "Preparado por:"
        ws["B5"].value = prepared_by

        for row in [4, 5]:
            ws[f"A{row}"].font = Font(bold=True)

    def _add_employee_info(self, ws, comparison: ComparisonResult,
                          current_company: str = "", current_position: str = "",
                          current_supervisor: str = "", proposal_company: str = "",
                          proposal_position: str = "", proposal_supervisor: str = ""):
        """Agrega información del empleado."""
        row = 7

        # Sección de datos del empleado
        self._add_section_header(ws, row, "INFORMACIÓN DEL COLABORADOR")
        row += 1

        employee_data = [
            ("Nombre Completo:", comparison.employee_name),
            ("RUT:", comparison.employee_rut),
            ("Fecha de Aplicación:", comparison.change_date),
        ]

        for label, value in employee_data:
            ws[f"A{row}"].value = label
            ws[f"B{row}"].value = value
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        # Agregar tabla de Empresa, Cargo, Jefe (Actual vs Propuesta)
        row += 1
        ws[f"A{row}"].value = "Item"
        ws[f"B{row}"].value = "Actual"
        ws[f"C{row}"].value = "Propuesta"
        for col in ["A", "B", "C"]:
            ws[f"{col}{row}"].font = Font(bold=True, color="FFFFFF")
            ws[f"{col}{row}"].fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")

        row += 1
        org_data = [
            ("Empresa", current_company or "N/A", proposal_company or current_company or "N/A"),
            ("Descripcion Cargo", current_position or "N/A", proposal_position or current_position or "N/A"),
            ("Nombre Jefe", current_supervisor or "N/A", proposal_supervisor or current_supervisor or "N/A"),
        ]

        for label, actual, proposal in org_data:
            ws[f"A{row}"].value = label
            ws[f"B{row}"].value = actual
            ws[f"C{row}"].value = proposal
            ws[f"A{row}"].font = Font(bold=True)
            for col in ["A", "B", "C"]:
                ws[f"{col}{row}"].alignment = Alignment(horizontal="left", vertical="center")
                ws[f"{col}{row}"].border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
            row += 1

        return row + 1

    def _add_comparison_table(self, ws, comparison: ComparisonResult):
        """Agrega tabla de comparación principal."""
        row = self._get_next_section_row(ws)

        self._add_section_header(ws, row, "COMPARATIVA DE HABERES Y DESCUENTOS")
        row += 1

        # Encabezados de tabla
        headers = ["Concepto", "Actual", "Propuesta", "Variación ($)", "Variación (%)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

        # Datos de comparación
        items = comparison.get_comparison_items()
        section_rows = {
            "HABERES": ["Sueldo Base", "Gratificación", "Colación", "Movilización", "Total Imponible", "Total No Imponible", "Total Haberes"],
            "DESCUENTOS": ["Descuento AFP", "Descuento Salud", "Descuento AFC", "Impuesto a la Renta", "Total Descuentos"],
            "RESULTADO NETO": ["Sueldo Líquido"],
            "COSTO EMPRESA": ["Costo Empresa"],
        }

        for section_name, concepts in section_rows.items():
            row = self._add_section_row(ws, row, section_name)

            for concept in concepts:
                if concept in items:
                    values = items[concept]
                    actual = values["actual"]
                    proposal = values["proposal"]
                    variation = proposal - actual
                    variation_pct = (variation / actual * 100) if actual != 0 else 0

                    ws.cell(row=row, column=1).value = concept
                    ws.cell(row=row, column=2).value = actual
                    ws.cell(row=row, column=3).value = proposal
                    ws.cell(row=row, column=4).value = variation
                    ws.cell(row=row, column=5).value = variation_pct / 100

                    # Formatear números y colores
                    for col in range(2, 6):
                        cell = ws.cell(row=row, column=col)
                        cell.number_format = "#,##0" if col < 5 else "0.00%"
                        cell.alignment = Alignment(horizontal="right")

                        if col == 4:  # Columna de variación
                            if variation > 0:
                                cell.fill = PatternFill(start_color=self.COLOR_POSITIVE, end_color=self.COLOR_POSITIVE, fill_type="solid")
                            elif variation < 0:
                                cell.fill = PatternFill(start_color=self.COLOR_NEGATIVE, end_color=self.COLOR_NEGATIVE, fill_type="solid")

                    row += 1

        return row

    def _add_impact_summary(self, ws, comparison: ComparisonResult):
        """Agrega resumen de impacto."""
        row = self._get_next_section_row(ws)

        self._add_section_header(ws, row, "RESUMEN DE IMPACTO")
        row += 1

        net_salary_current = comparison.current.net_salary
        net_salary_proposal = comparison.proposal.net_salary
        net_salary_impact = net_salary_proposal - net_salary_current

        cost_employer_current = comparison.current.total_employer_cost
        cost_employer_proposal = comparison.proposal.total_employer_cost
        cost_employer_impact = cost_employer_proposal - cost_employer_current

        impact_data = [
            ("Impacto en Sueldo Líquido (Mensual):", net_salary_impact),
            ("Impacto en Costo Empresa (Mensual):", cost_employer_impact),
            ("Impacto Total Nómina (Mensual):", net_salary_impact + cost_employer_impact),
        ]

        for label, value in impact_data:
            ws[f"A{row}"].value = label
            ws[f"B{row}"].value = value
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].number_format = "#,##0"
            ws[f"B{row}"].alignment = Alignment(horizontal="right")

            if value > 0:
                ws[f"B{row}"].fill = PatternFill(start_color=self.COLOR_POSITIVE, end_color=self.COLOR_POSITIVE, fill_type="solid")
            elif value < 0:
                ws[f"B{row}"].fill = PatternFill(start_color=self.COLOR_NEGATIVE, end_color=self.COLOR_NEGATIVE, fill_type="solid")

            row += 1

        return row + 1

    def _add_observations(self, ws, proposal_reasons: list = None):
        """Agrega sección de observaciones."""
        row = self._get_next_section_row(ws)

        self._add_section_header(ws, row, "OBSERVACIONES")
        row += 1

        # Agregar motivos de la propuesta si existen
        if proposal_reasons:
            ws[f"A{row}"].value = "Motivos de la Propuesta:"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            for reason in proposal_reasons:
                ws[f"A{row}"].value = f"• {reason}"
                row += 1

            row += 1

        ws.merge_cells(f"A{row}:E{row + 4}")
        obs_cell = ws[f"A{row}"]
        obs_cell.value = "- Esta propuesta debe ser aprobada por el departamento de Recursos Humanos y Finanzas.\n- Los valores pueden variar según actualizaciones de parámetros mensuales (UF, UTM, IMM).\n- Consulte con el equipo de Nómina ante dudas sobre los cálculos."
        obs_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 60

    def _add_section_header(self, ws, row: int, title: str):
        """Agrega encabezado de sección."""
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color=self.COLOR_SECTION, end_color=self.COLOR_SECTION, fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18

    def _add_section_row(self, ws, row: int, section_name: str) -> int:
        """Agrega fila de subtítulo de sección con color."""
        ws.merge_cells(f"A{row}:E{row}")
        cell = ws[f"A{row}"]
        cell.value = section_name
        cell.font = Font(bold=True, italic=True)
        cell.fill = PatternFill(start_color=self.COLOR_TOTAL, end_color=self.COLOR_TOTAL, fill_type="solid")
        ws.row_dimensions[row].height = 16
        return row + 1

    def _get_next_section_row(self, ws) -> int:
        """Obtiene siguiente fila disponible con espaciado."""
        for row in range(ws.max_row, 0, -1):
            if ws[f"A{row}"].value is not None:
                return row + 2
        return 1

    def _add_salary_history(self, ws, salary_history: list):
        """Agrega historial de sueldos en una hoja."""
        # Configurar ancho de columnas
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 18

        row = 1

        # Título
        ws.merge_cells("A1:D1")
        title = ws["A1"]
        title.value = "HISTORIAL DE SUELDOS"
        title.font = Font(bold=True, size=14, color="FFFFFF")
        title.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 25

        row = 3

        # Encabezados
        headers = ["Periodo", "Sueldo Base", "Variación ($)", "Variación (%)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row += 1

        # Filtrar registros: excluir mayo 2019 y anteriores
        filtered_history = [record for record in salary_history if record.get("start_date", "")[:7] > "2019-05"]

        # Procesar datos
        for i, record in enumerate(filtered_history):
            start = record.get("start_date", "")
            wage = record.get("base_wage", 0)

            # Extraer solo mes y año
            periodo = start[:7] if start else "N/A"

            # Calcular variación
            variation = 0
            variation_pct = 0

            if i < len(filtered_history) - 1:  # Si no es el último (más antiguo)
                prev_wage = filtered_history[i + 1].get("base_wage", 0)
                if prev_wage:
                    variation = wage - prev_wage
                    variation_pct = (variation / prev_wage * 100) if prev_wage > 0 else 0

            # Agregar fila
            ws.cell(row=row, column=1).value = periodo
            ws.cell(row=row, column=2).value = wage
            ws.cell(row=row, column=3).value = variation
            ws.cell(row=row, column=4).value = variation_pct / 100

            # Formatear
            ws.cell(row=row, column=2).number_format = "$#,##0"
            ws.cell(row=row, column=3).number_format = "$#,##0"
            ws.cell(row=row, column=4).number_format = "0.00%"

            for col in range(1, 5):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")

            # Colorear variación
            if variation > 0:
                ws.cell(row=row, column=3).fill = PatternFill(start_color=self.COLOR_POSITIVE, end_color=self.COLOR_POSITIVE, fill_type="solid")
            elif variation < 0:
                ws.cell(row=row, column=3).fill = PatternFill(start_color=self.COLOR_NEGATIVE, end_color=self.COLOR_NEGATIVE, fill_type="solid")

            row += 1

        # Resumen
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        summary = ws[f"A{row}"]
        summary.value = "RESUMEN"
        summary.font = Font(bold=True, color="FFFFFF")
        summary.fill = PatternFill(start_color=self.COLOR_SECTION, end_color=self.COLOR_SECTION, fill_type="solid")
        summary.alignment = Alignment(horizontal="left", vertical="center")

        row += 1

        if len(filtered_history) > 0:
            # Total de períodos
            ws[f"A{row}"].value = "Total Periodos:"
            ws[f"B{row}"].value = len(filtered_history)
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

            # Aumento total
            if len(filtered_history) > 1:
                first_wage = filtered_history[-1].get("base_wage", 0)
                current_wage = filtered_history[0].get("base_wage", 0)
                total_increase = current_wage - first_wage
                total_increase_pct = (total_increase / first_wage * 100) if first_wage > 0 else 0

                ws[f"A{row}"].value = "Aumento Total:"
                ws[f"B{row}"].value = total_increase
                ws[f"C{row}"].value = total_increase_pct / 100
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"].number_format = "$#,##0"
                ws[f"C{row}"].number_format = "0.00%"
                row += 1

            # Sueldo inicial
            first = filtered_history[-1].get("base_wage", 0)
            ws[f"A{row}"].value = "Sueldo Inicial:"
            ws[f"B{row}"].value = first
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"].number_format = "$#,##0"

    def export_calculator(self, calculation, output_filename: str, periodo: str = "") -> bool:
        """
        Exporta resultado de calculadora a Excel.

        Args:
            calculation: PayrollCalculation object
            output_filename: Ruta del archivo Excel
            periodo: Período a mostrar

        Returns:
            True si fue exitoso, False en caso contrario
        """
        try:
            self.workbook = Workbook()
            ws = self.workbook.active
            ws.title = "Liquidación"

            # Configurar ancho de columnas
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 18

            row = 1

            # Título
            ws.merge_cells(f"A{row}:C{row}")
            title_cell = ws[f"A{row}"]
            title_cell.value = "LIQUIDACIÓN DE SUELDO"
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 25

            row += 2

            # Período
            if periodo:
                ws[f"A{row}"] = "Período:"
                ws[f"B{row}"] = periodo
                row += 1

            row += 1

            # HABERES IMPONIBLES
            ws.merge_cells(f"A{row}:C{row}")
            header = ws[f"A{row}"]
            header.value = "HABERES IMPONIBLES"
            header.font = Font(bold=True, color="FFFFFF")
            header.fill = PatternFill(start_color=self.COLOR_SECTION, end_color=self.COLOR_SECTION, fill_type="solid")
            row += 1

            haberes = [
                ("Sueldo Base", calculation.base_salary),
                ("Gratificación", calculation.gratification),
                ("Colación", calculation.collation),
                ("Movilización", calculation.mobility),
                ("Otros Imponibles", calculation.other_taxable),
            ]

            for concepto, valor in haberes:
                ws[f"A{row}"] = concepto
                ws[f"B{row}"] = valor
                ws[f"B{row}"].number_format = '$#,##0'
                row += 1

            # Total imponible
            ws[f"A{row}"] = "TOTAL IMPONIBLE"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = calculation.total_taxable
            ws[f"B{row}"].font = Font(bold=True)
            ws[f"B{row}"].number_format = '$#,##0'
            ws[f"B{row}"].fill = PatternFill(start_color=self.COLOR_TOTAL, end_color=self.COLOR_TOTAL, fill_type="solid")

            row += 2

            # HABERES NO IMPONIBLES
            ws.merge_cells(f"A{row}:C{row}")
            header = ws[f"A{row}"]
            header.value = "HABERES NO IMPONIBLES"
            header.font = Font(bold=True, color="FFFFFF")
            header.fill = PatternFill(start_color=self.COLOR_SECTION, end_color=self.COLOR_SECTION, fill_type="solid")
            row += 1

            ws[f"A{row}"] = "Total No Imponible"
            ws[f"B{row}"] = calculation.total_non_taxable
            ws[f"B{row}"].number_format = '$#,##0'
            row += 2

            # DESCUENTOS LEGALES
            ws.merge_cells(f"A{row}:C{row}")
            header = ws[f"A{row}"]
            header.value = "DESCUENTOS LEGALES"
            header.font = Font(bold=True, color="FFFFFF")
            header.fill = PatternFill(start_color=self.COLOR_SECTION, end_color=self.COLOR_SECTION, fill_type="solid")
            row += 1

            descuentos = [
                ("AFP", calculation.afp_discount),
                ("Salud", calculation.health_discount),
                ("AFC (Seguro de Cesantía)", calculation.afc_discount),
                ("Impuesto a la Renta", calculation.income_tax),
            ]

            for concepto, valor in descuentos:
                ws[f"A{row}"] = concepto
                ws[f"B{row}"] = valor
                ws[f"B{row}"].number_format = '$#,##0'
                row += 1

            # Total descuentos
            ws[f"A{row}"] = "TOTAL DESCUENTOS"
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"B{row}"] = calculation.total_discounts
            ws[f"B{row}"].font = Font(bold=True)
            ws[f"B{row}"].number_format = '$#,##0'
            ws[f"B{row}"].fill = PatternFill(start_color=self.COLOR_TOTAL, end_color=self.COLOR_TOTAL, fill_type="solid")

            row += 2

            # RESUMEN FINAL
            ws.merge_cells(f"A{row}:C{row}")
            header = ws[f"A{row}"]
            header.value = "RESUMEN FINAL"
            header.font = Font(bold=True, color="FFFFFF")
            header.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            row += 1

            resumen = [
                ("Total Haberes", calculation.total_earnings),
                ("Total Descuentos", calculation.total_discounts),
                ("SUELDO LÍQUIDO", calculation.net_salary),
            ]

            for concepto, valor in resumen:
                ws[f"A{row}"] = concepto
                ws[f"A{row}"].font = Font(bold=True)
                ws[f"B{row}"] = valor
                ws[f"B{row}"].font = Font(bold=True)
                ws[f"B{row}"].number_format = '$#,##0'
                if concepto == "SUELDO LÍQUIDO":
                    ws[f"B{row}"].fill = PatternFill(start_color=self.COLOR_POSITIVE, end_color=self.COLOR_POSITIVE, fill_type="solid")
                row += 1

            # Guardar archivo
            self.workbook.save(output_filename)
            return True

        except Exception as e:
            logger.error(f"Error exportando calculadora a Excel: {e}")
            return False

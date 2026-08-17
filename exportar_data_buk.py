"""Script para exportar TODOS los datos de Buk a Excel"""
import os
import json
from dotenv import load_dotenv
from src.buk_client import BukClient
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

load_dotenv()

buk = BukClient(os.getenv('BUK_API_TOKEN'), os.getenv('BUK_SUBDOMAIN'))

ruts = ["17771319-8", "13531650-4"]

# Crear workbook
wb = Workbook()
wb.remove(wb.active)

print("Extrayendo datos de Buk...")

for rut in ruts:
    print(f"\n  Procesando RUT: {rut}")

    try:
        # Buscar empleado
        employee = buk.search_employee(rut=rut)

        if not employee:
            print(f"    ✗ Empleado no encontrado")
            continue

        print(f"    ✓ {employee.full_name}")

        # ===== HOJA 1: DATOS BÁSICOS Y VIGENTES =====
        ws_vigente = wb.create_sheet(f"{rut} - Vigente")

        # Extraer atributos disponibles del employee
        employee_attrs = [attr for attr in dir(employee) if not attr.startswith('_')]

        row = 1
        for attr in employee_attrs:
            try:
                value = getattr(employee, attr)
                # Saltar métodos y objetos complejos
                if callable(value) or isinstance(value, (dict, list)) and len(str(value)) > 100:
                    continue

                ws_vigente[f"A{row}"] = attr
                ws_vigente[f"B{row}"] = str(value) if value else ""
                ws_vigente[f"A{row}"].font = Font(bold=True)
                row += 1
            except:
                pass

        # Current Job (Ficha Vigente)
        row += 1
        ws_vigente[f"A{row}"] = "FICHA VIGENTE (Current Job)"
        ws_vigente[f"A{row}"].font = Font(bold=True, color="FFFFFF")
        ws_vigente[f"A{row}"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        row += 1

        if employee.current_job:
            for key, value in employee.current_job.items():
                ws_vigente[f"A{row}"] = key
                ws_vigente[f"B{row}"] = str(value) if value else ""
                ws_vigente[f"A{row}"].font = Font(bold=True)
                row += 1
        else:
            ws_vigente[f"A{row}"] = "⚠️ Sin ficha vigente"
            row += 1

        # Ajustar ancho de columnas
        ws_vigente.column_dimensions["A"].width = 30
        ws_vigente.column_dimensions["B"].width = 50

        # ===== HOJA 2: HISTORIAL DE PUESTOS =====
        print("    → Extrayendo historial de puestos...")
        jobs = buk.get_job_history(rut)
        if jobs and len(jobs) > 0:
            ws_jobs = wb.create_sheet(f"{rut} - Puestos")

            # Headers
            headers = list(jobs[0].keys()) if jobs else []
            for col, header in enumerate(headers, 1):
                cell = ws_jobs.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

            # Datos
            for row, job in enumerate(jobs, 2):
                for col, header in enumerate(headers, 1):
                    cell = ws_jobs.cell(row=row, column=col)
                    value = job.get(header)
                    cell.value = str(value) if value else ""
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Ajustar ancho
            for col in range(1, len(headers) + 1):
                ws_jobs.column_dimensions[chr(64 + col)].width = 18

            print(f"      ✓ {len(jobs)} puestos encontrados")

        # ===== HOJA 3: HISTORIAL DE SUELDOS =====
        print("    → Extrayendo historial de sueldos...")
        salary_hist = buk.get_salary_history(rut)
        if salary_hist and len(salary_hist) > 0:
            ws_salary = wb.create_sheet(f"{rut} - Sueldos")

            # Headers
            headers = list(salary_hist[0].keys()) if salary_hist else []
            for col, header in enumerate(headers, 1):
                cell = ws_salary.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

            # Datos
            for row, sal in enumerate(salary_hist, 2):
                for col, header in enumerate(headers, 1):
                    cell = ws_salary.cell(row=row, column=col)
                    value = sal.get(header)
                    cell.value = str(value) if value else ""
                    # Formatear números
                    if "wage" in str(header).lower() or "salary" in str(header).lower():
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Ajustar ancho
            for col in range(1, len(headers) + 1):
                ws_salary.column_dimensions[chr(64 + col)].width = 18

            print(f"      ✓ {len(salary_hist)} períodos de sueldo encontrados")

    except Exception as e:
        print(f"    ✗ Error: {str(e)}")
        import traceback
        traceback.print_exc()

# Guardar archivo
filename = f"Data_Buk_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(filename)

print(f"\n✅ Archivo guardado: {filename}")
print(f"   📁 Ubicación: {os.path.abspath(filename)}")

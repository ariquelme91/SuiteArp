"""Script avanzado para extraer datos de Buk incluyendo Haberes y Descuentos por ID"""
import os
import json
import requests
import re
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

load_dotenv()

token = os.getenv('BUK_API_TOKEN')
subdomain = os.getenv('BUK_SUBDOMAIN')

base_url = f"https://{subdomain}.buk.cl/api/v1/chile"
headers = {
    "auth_token": token,
    "Content-Type": "application/json"
}

ruts = ["17771319-8", "13531650-4", "18.732.184-0"]

def format_rut(rut):
    return re.sub(r'[.-]', '', rut).upper()

def obtener_bono(bono_id):
    """Obtiene detalles completos de un bono/haber"""
    try:
        response = requests.get(
            f"{base_url}/bonos/{bono_id}",
            headers=headers,
            timeout=10
        )
        if response.status_code == 200:
            data = response.json()
            return data.get("data", data)
    except:
        pass
    return None

def obtener_descuento(descuento_id):
    """Obtiene detalles completos de un descuento"""
    try:
        response = requests.get(
            f"{base_url}/descuentos/{descuento_id}",
            headers=headers,
            timeout=10
        )
        if response.status_code == 200:
            data = response.json()
            return data.get("data", data)
    except:
        pass
    return None

def add_header_style(ws, row):
    """Agrega estilo a los headers"""
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

print("Extrayendo datos avanzados de Buk (incluyendo Haberes y Descuentos)...")

wb = Workbook()
wb.remove(wb.active)

for rut in ruts:
    print(f"\nProcesando: {rut}")

    try:
        rut_formatted = format_rut(rut)
        response = requests.get(
            f"{base_url}/employees/{rut_formatted}",
            headers=headers,
            timeout=10
        )
        response.raise_for_status()

        raw_data = response.json()
        employee_data = raw_data.get("data", raw_data)

        # Verificar que sea ficha vigente/activa
        status = employee_data.get("status", "")
        current_job = employee_data.get("current_job", {})
        job_status = current_job.get("status", "")

        if status != "activo" and job_status != "activo":
            print(f"  SALTADO - No es ficha activa")
            continue

        nombre = employee_data.get("full_name", rut)
        print(f"  OK - {nombre}")

        # ===== HOJA 1: DATOS BÁSICOS =====
        ws_basicos = wb.create_sheet(f"{rut} - Basicos")

        row = 1
        datos_basicos = {
            "RUT": employee_data.get("rut"),
            "Nombre": employee_data.get("full_name"),
            "Email": employee_data.get("email"),
            "Cargo": current_job.get("name"),
            "Empresa": current_job.get("company_name"),
            "Estado": current_job.get("status"),
            "Sueldo Base": current_job.get("base_wage"),
            "Fecha Ingreso": current_job.get("start_date"),
            "Contrato": current_job.get("contract_type"),
            "Fondo Pensión": employee_data.get("pension_fund"),
        }

        for key, value in datos_basicos.items():
            ws_basicos[f"A{row}"] = key
            ws_basicos[f"B{row}"] = value
            ws_basicos[f"A{row}"].font = Font(bold=True)
            row += 1

        ws_basicos.column_dimensions["A"].width = 25
        ws_basicos.column_dimensions["B"].width = 50

        # ===== HOJA 2: HABERES (BONOS) =====
        print("    → Extrayendo haberes (bonos)...")
        fixed_items = current_job.get("fixed_items", [])
        haberes_detalle = []

        for item in fixed_items:
            item_id = item.get("id")
            if item_id:
                bono = obtener_bono(item_id)
                if bono:
                    haberes_detalle.append(bono)
                    print(f"      ✓ {bono.get('name', 'Sin nombre')}")

        if haberes_detalle:
            ws_haberes = wb.create_sheet(f"{rut} - Haberes")

            # Headers dinámicos basados en el primer haber
            if haberes_detalle:
                headers = list(haberes_detalle[0].keys())
                for col, header in enumerate(headers, 1):
                    ws_haberes.cell(row=1, column=col).value = header

                add_header_style(ws_haberes, 1)

                # Datos
                for row, haber in enumerate(haberes_detalle, 2):
                    for col, header in enumerate(headers, 1):
                        value = haber.get(header)
                        ws_haberes.cell(row=row, column=col).value = str(value) if value is not None else ""

                # Ajustar ancho
                for col in range(1, len(headers) + 1):
                    ws_haberes.column_dimensions[chr(64 + col)].width = 18

        # ===== HOJA 3: DESCUENTOS =====
        print("    → Extrayendo descuentos...")
        discounts_raw = current_job.get("discounts", [])
        descuentos_detalle = []

        for discount in discounts_raw:
            discount_id = discount.get("id")
            if discount_id:
                desc = obtener_descuento(discount_id)
                if desc:
                    descuentos_detalle.append(desc)
                    print(f"      ✓ {desc.get('name', 'Sin nombre')}")

        if descuentos_detalle:
            ws_descuentos = wb.create_sheet(f"{rut} - Descuentos")

            # Headers dinámicos
            if descuentos_detalle:
                headers = list(descuentos_detalle[0].keys())
                for col, header in enumerate(headers, 1):
                    ws_descuentos.cell(row=1, column=col).value = header

                add_header_style(ws_descuentos, 1)

                # Datos
                for row, desc in enumerate(descuentos_detalle, 2):
                    for col, header in enumerate(headers, 1):
                        value = desc.get(header)
                        ws_descuentos.cell(row=row, column=col).value = str(value) if value is not None else ""

                # Ajustar ancho
                for col in range(1, len(headers) + 1):
                    ws_descuentos.column_dimensions[chr(64 + col)].width = 18

    except Exception as e:
        print(f"  ERROR: {str(e)}")
        import traceback
        traceback.print_exc()

# Guardar archivo Excel
filename = f"Data_Buk_Avanzado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(filename)

print(f"\n✅ Archivo Excel creado: {filename}")
print(f"   Ubicacion: C:\\Users\\ariquelme\\propuestas de renta\\{filename}")

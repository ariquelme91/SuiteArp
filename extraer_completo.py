"""Script para extraer TODOS los datos de Buk y guardarlos en Excel"""
import os
import json
import requests
import re
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

load_dotenv()

token = os.getenv('BUK_API_TOKEN')
subdomain = os.getenv('BUK_SUBDOMAIN')

base_url = f"https://{subdomain}.buk.cl/api/v1/chile"
headers = {
    "auth_token": token,
    "Content-Type": "application/json"
}

ruts = ["17771319-8", "13531650-4", "18.732.184-0"]

def format_rut(rut):
    return re.sub(r'[.-]', '', rut).upper()

def flatten_dict(d, parent_key='', sep='_'):
    """Convierte diccionarios anidados en columnas planas"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            items.append((new_key, json.dumps(v, default=str)))
        else:
            items.append((new_key, v))
    return dict(items)

print("Extrayendo datos de Buk...")

wb = Workbook()
wb.remove(wb.active)

all_data = {}

for rut in ruts:
    print(f"\nProcesando: {rut}")

    try:
        rut_formatted = format_rut(rut)
        response = requests.get(
            f"{base_url}/employees/{rut_formatted}",
            headers=headers,
            timeout=10
        )
        response.raise_for_status()

        raw_data = response.json()

        # Extraer el objeto employee
        if "data" in raw_data:
            employee_data = raw_data["data"]
        else:
            employee_data = raw_data

        all_data[rut] = employee_data

        # Crear hoja para este RUT
        ws = wb.create_sheet(rut)

        # APLANAR TODO (incluyendo atributos personalizados)
        flat_data = flatten_dict(employee_data)

        # Escribir en columnas A y B
        row = 1
        for key, value in flat_data.items():
            ws[f"A{row}"] = key
            ws[f"B{row}"] = str(value) if value is not None else ""
            ws[f"A{row}"].font = Font(bold=True)
            row += 1

        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 60

        print(f"  OK - {len(flat_data)} campos encontrados")

    except Exception as e:
        print(f"  ERROR: {str(e)}")

# Guardar archivo Excel
filename = f"Data_Buk_Completa_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(filename)

print(f"\nArchivo Excel creado: {filename}")

# También guardar JSON para referencia
json_filename = f"Data_Buk_Raw_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
with open(json_filename, 'w', encoding='utf-8') as f:
    json.dump(all_data, f, indent=2, default=str, ensure_ascii=False)

print(f"Archivo JSON creado: {json_filename}")
